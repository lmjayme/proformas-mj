import streamlit as st
import openpyxl
from openpyxl.drawing.image import Image
import os

st.set_page_config(page_title="MJ LOGISTIC - Editor Pro", layout="wide")
st.title("💼 Generador de Proformas - MJ LOGISTIC")

# 1. SUBIDA DEL ARCHIVO DESDE TU CELULAR
archivo_subido = st.file_uploader("📂 Selecciona tu archivo Excel (.xlsx):", type=["xlsx"])
archivo_logo = "logo.png"

if archivo_subido:
    wb = openpyxl.load_workbook(archivo_subido)
    ws = wb.active
    
    with st.expander("📝 Editar Datos de la Proforma", expanded=True):
        col1, col2, col3 = st.columns(3)
        nro = col1.text_input("Nro Proforma", "MJ240114")
        cliente = col1.text_input("Cliente", "ENERGY AND SOLUTIONS ELECTRICAL SAC")
        fecha = col2.text_input("Fecha", "2026-06-07")
        atencion = col2.text_input("Atención", "Srta.")
        servicio = col3.text_input("Servicio", "LCL MARITIMO")
        proveedor = col3.text_input("Proveedor", "-")

    # 2. GASTOS DINÁMICOS
    st.subheader("➕ Gastos y Conceptos")
    if 'gastos' not in st.session_state:
        st.session_state.gastos = [{"concepto": "DESCARGA", "monto": 35.4, "ref": "MIN. 30+IGV"}, {"concepto": "V°B°", "monto": 118.0, "ref": "100+IGV"}]

    for i, g in enumerate(st.session_state.gastos):
        cols = st.columns([3, 1, 1, 1])
        st.session_state.gastos[i]['concepto'] = cols[0].text_input(f"Concepto {i+1}", g['concepto'], key=f"c{i}")
        st.session_state.gastos[i]['monto'] = cols[1].number_input(f"Monto {i+1}", value=float(g['monto']), key=f"m{i}")
        st.session_state.gastos[i]['ref'] = cols[2].text_input(f"Ref {i+1}", g['ref'], key=f"r{i}")
        if cols[3].button("🗑️", key=f"d{i}"):
            st.session_state.gastos.pop(i)
            st.rerun()

    if st.button("➕ Agregar Fila"):
        st.session_state.gastos.append({"concepto": "", "monto": 0.0, "ref": ""})
        st.rerun()

    # 3. GENERACIÓN FINAL (Inyección + Nombre Personalizado)
    if st.button("🚀 Preparar Archivo para Guardar"):
        # Insertar Logo si existe
        if os.path.exists(archivo_logo):
            img = Image(archivo_logo)
            img.width = 120 
            img.height = 40
            ws.add_image(img, 'A1')
            
        # Inyección de datos
        ws['A4'] = f"PROFORMA {nro}"
        ws['A9'] = cliente
        ws['F5'] = fecha
        
        # Gastos dinámicos (fila 20)
        fila = 20
        for g in st.session_state.gastos:
            ws.cell(row=fila, column=9, value=g['concepto'])
            ws.cell(row=fila, column=10, value=g['monto'])
            ws.cell(row=fila, column=12, value=g['ref'])
            fila += 1
            
        # Guardar en memoria temporal
        archivo_temp = "temporal.xlsx"
        wb.save(archivo_temp)
        
        # GUARDAR ESTADO PARA QUE APAREZCA EL INPUT DE NOMBRE
        st.session_state.archivo_generado = archivo_temp
        st.success("¡Datos inyectados! Ahora ponle nombre y descarga.")

        #4. EDITAR NOMBRE Y DESCARGAR
        if 'archivo_generado' in st.session_state:
        nombre_final = st.text_input("💾 Nombre del archivo (sin .xlsx):", f"PROFORMA_{nro}")
        
        with open(st.session_state.archivo_generado, "rb") as f:
            st.download_button(
                label=f"📥 Descargar {nombre_final}.xlsx",
                data=f,
                file_name=f"{nombre_final}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


    
